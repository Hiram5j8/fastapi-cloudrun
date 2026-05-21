from fastapi import FastAPI, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse

import os
import shutil
import zipfile
import tempfile
import glob

from openpyxl import Workbook, load_workbook

app = FastAPI()

app.mount("/static", StaticFiles(directory="static", html=True), name="static")


@app.get("/")
def home():
    return FileResponse("static/index.html")


@app.post("/process")
async def process(data_zip: UploadFile = File(...)):

    # =========================
    # 建立工作目錄
    # =========================
    work = tempfile.mkdtemp(prefix="work_")

    # ZIP
    zip_path = os.path.join(work, "data.zip")

    with open(zip_path, "wb") as f:
        shutil.copyfileobj(data_zip.file, f)

    # =========================
    # 解壓 ZIP
    # =========================
    unzip_dir = os.path.join(work, "unzipped")
    os.makedirs(unzip_dir, exist_ok=True)

    with zipfile.ZipFile(zip_path, "r") as z:
        z.extractall(unzip_dir)

    # =========================
    # 找全部 Excel
    # =========================
    files = glob.glob(
        os.path.join(unzip_dir, "**", "*.xls*"),
        recursive=True
    )

    if not files:
        return {"error": "ZIP 沒有 Excel"}

    # =========================
    # 建立輸出 Workbook
    # =========================
    out_wb = Workbook()

    # 刪除預設 Sheet
    out_wb.remove(out_wb.active)

    # =========================
    # 開始合併
    # =========================
    for excel_file in files:

        try:
            wb = load_workbook(excel_file, data_only=True)

            for sheet_name in wb.sheetnames:

                src_ws = wb[sheet_name]

                # =========================
                # 如果 sheet 不存在 → 建立
                # =========================
                if sheet_name not in out_wb.sheetnames:

                    out_ws = out_wb.create_sheet(sheet_name)

                    # 完整複製
                    for row in src_ws.iter_rows():

                        for cell in row:

                            out_ws[cell.coordinate] = cell.value

                else:

                    # =========================
                    # sheet 已存在 → 接在下面
                    # =========================
                    out_ws = out_wb[sheet_name]

                    start_row = out_ws.max_row + 1

                    for row in src_ws.iter_rows():

                        for cell in row:

                            new_row = cell.row + start_row - 1

                            out_ws.cell(
                                row=new_row,
                                column=cell.column
                            ).value = cell.value

        except Exception as e:
            print("錯誤:", excel_file, e)

    # =========================
    # 儲存
    # =========================
    result_path = os.path.join(work, "result.xlsx")

    out_wb.save(result_path)

    # =========================
    # 回傳
    # =========================
    return FileResponse(
        result_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="result.xlsx"
    )