import os
import sys
import glob
import shutil
import re
import unicodedata
import math
import hashlib
import numpy as np
import pandas as pd
import xlwings as xw
import tkinter as tk
import zipfile, tempfile

from datetime import datetime
from tkinter import filedialog, messagebox
from openpyxl import load_workbook


# ===============================
# 🔥 授權（可選）
# ===============================
root = tk.Tk()
root.withdraw()





# 正常執行程式
print("🎉 程式執行中...")

# ===============================
# 🔥 正規化（100%防假空白）
# ===============================
def norm(v):
    if v is None:
        return ""

    if isinstance(v, float) and math.isnan(v):
        return ""

    if isinstance(v, float) and v.is_integer():
        v = int(v)

    v = str(v)

    v = unicodedata.normalize("NFKC", v)

    v = (
        v.replace("\xa0", "")
         .replace("\u3000", "")
         .replace("\ufeff", "")
         .replace("\u200b", "")
         .replace("\u200c", "")
         .replace("\u200d", "")
         .replace("\u00ad", "")
    )

    v = re.sub(r"[\x00-\x1F\x7F]", "", v)
    v = re.sub(r"\s+", "", v)

    return v.lower()


def key(v):
    return hashlib.md5(norm(v).encode("utf-8")).hexdigest()


# ===============================
# 🔥 safe dataframe access
# ===============================
def safe(df, r, c):
    if r >= len(df):
        return None
    if c >= df.shape[1]:
        return None
    return df.iat[r, c]


# ===============================
# 🔥 report
# ===============================
report = []

def add_report(sheet, source, master, row):
    report.append({
        "sheet": sheet,
        "source": source,
        "master": master,
        "row": row
    })


def export_report(path):
    pd.DataFrame(report).to_excel(path, index=False)


# ===============================
# 🔥 file utils
# ===============================
def get_files(folder):
    return [
        f for f in os.listdir(folder)
        if f.endswith((".xls", ".xlsx", ".xlsm")) and not f.startswith("~$")
    ]


def pick_folder(title):
    """  """
    max = 2
    cancel_count = 0
    while True:
        f = filedialog.askdirectory(title=title)
        cancel_count +=1 
        if f:
            return os.path.abspath(f)
        messagebox.showwarning("必要選擇", "一定要選擇資料夾")
        if cancel_count >= max:           
            sys.exit()
            

def pick_file(folder):
    files = get_files(folder)
    if len(files) == 1:
        return os.path.join(folder, files[0])

    return filedialog.askopenfilename(initialdir=folder)


# ===============================
# 🔥 open excel
# ===============================
def open_excel(path):
    ext = os.path.splitext(path)[1].lower()

    if ext in [".xlsx", ".xlsm"]:
        return ("openpyxl", load_workbook(path))

    return ("pandas", pd.ExcelFile(path))


# ===============================
# 🔥 master index（核心）
# ===============================
def build_index(ws):
    idx = {}

    for r in range(3, 153, 5):
        v = ws.range(r, 1).value
        k = key(v)

        if norm(v) != "":
            idx[k] = r

    return idx


def match(index, v):
    return index.get(key(v))

def pick_zip_file(title="選 ZIP 檔"):
    path = filedialog.askopenfilename(
        title=title,
        filetypes=[("ZIP files", "*.zip")]
    )
    return path


# ===============================
# 🚀 主程式
# ===============================
def main():

    master_folder = pick_folder("選 Master")
    master_file = pick_file(master_folder)

    # === 選 ZIP ===
    zip_path = pick_zip_file("選 Source ZIP")

    # === 解壓 ===
    unzip_dir = tempfile.mkdtemp(prefix="unzipped_")
    with zipfile.ZipFile(zip_path, 'r') as z:
        z.extractall(unzip_dir)

    files = glob.glob(os.path.join(unzip_dir, "**", "*.xls*"), recursive=True)

    # === 複製 master ===
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    new_id, ext = os.path.splitext(master_file)
    out_file = os.path.join(master_folder, f"{new_id}_{ts}.xlsx")
    shutil.copy(master_file, out_file)

    # === 開 Excel ===
    app = xw.App(visible=False)
    wb = xw.Book(out_file)

    engine, master_wb = open_excel(master_file)

    # 之後你原本對 files 的處理流程完全不用改
    for f in files:
        print("處理:", f)

    sheets = (
        master_wb.sheetnames
        if engine == "openpyxl"
        else master_wb.sheet_names
    )

    print("Sheets:", sheets)

    LEAVE = {"休", "請假", "休假", "特休", "請假一天"}
   
    # ===============================
    # 🔥 主處理
    # ===============================
    for file in files:

        try:
            data = pd.read_excel(file, sheet_name=None)
        except:
            continue

        for sh in sheets:

            if sh not in data:
                continue

            if sh not in [s.name for s in wb.sheets]:
                continue

            ws = wb.sheets[sh]
            df = data[sh]

            if df.shape[1] < 2:
                continue

            index = build_index(ws)

            for start in range(3, 153, 5):

                r0 = start - 2

                if r0 >= len(df):
                    break

                src = norm(df.iloc[r0, 0])

                if src == "":
                    continue

                found = match(index, src)

                # =========================
                # 不存在 → 新增
                # =========================
                if found is None:

                    add_report(sh, src, "", -1)

                    for r in range(3, 153, 5):
                        if norm(ws.range(r, 1).value) == "":
                            found = r
                            ws.range(r, 1).value = src
                            index[key(src)] = r
                            break

                    if found is None:
                        continue

                # =========================
                # 寫入 B 欄（只寫空白 & 非公式）
                # =========================
                for i in range(5):

                    r = start + i - 2
                    val = safe(df, r, 1)

                    if val is not None and pd.notna(val):
                        ws.range(found + i, 2).value = str(val)
                # =========================
                # C ~ MAX（完全自動）
                # =========================
                #max_c = df.shape[1]
                max_c = 37		#AK
                
                for i in range(5):

                    r = start + i - 2

                    for c in range(2, max_c):

                        val = safe(df, r, c)

                        if val is None or pd.isna(val):
                            continue

                        if isinstance(val, (int, float, np.number)):
                            ws.range(found + i, c + 1).value = val

                        elif isinstance(val, str) and val.strip() in LEAVE:
                            ws.range(found + i, c + 1).value = "請假"

    # ===============================
    # save
    # ===============================
    wb.save(out_file)
    wb.close()
    app.quit()

    #export_report(os.path.join(master_folder, "report.xlsx"))
    print("完成:", out_file)
    messagebox.showinfo("提示", "處理完成！")


if __name__ == "__main__":
    main()
