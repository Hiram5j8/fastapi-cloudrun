from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook, Workbook
from dataclasses import dataclass
from typing import Any, List
from io import BytesIO

import random
import zipfile
import os
import shutil
import tempfile
import re

# =====================================================
# FastAPI
# =====================================================

app = FastAPI()

#global current_captcha

# ✅ static
app.mount(
    "/static",
    StaticFiles(directory="static", html=True),
    name="static"
)

# ✅ 首頁
@app.get("/")
def home():
    return FileResponse("static/index.html")

# =========================
# 版本 API
# =========================
VERSION = "2026-05-26 1022"

@app.get("/version")
def version():
    return {
        "version": VERSION
    }

# =========================
# CAPTCHA API
# =========================
@app.get("/captcha")
def get_captcha():

    global current_captcha

    current_captcha = str(
        random.randint(100000, 999999)
    )
    return {
        "captcha": current_captcha
    }
# =========================
# 比較差異
# =========================
def compare_sheet(base_ws, src_ws):

    diff_rows = []

    max_row = min(base_ws.max_row, src_ws.max_row)
    max_col = min(base_ws.max_column, src_ws.max_column)

    for r in range(1, max_row + 1):

        row_values = []
        has_diff = False

        for c in range(1, max_col + 1):

            base_val = base_ws.cell(r, c).value
            src_val = src_ws.cell(r, c).value

            # 差異判斷
            if base_val != src_val:
                has_diff = True

            row_values.append(src_val)

        # 有差異才輸出
        if has_diff:
            diff_rows.append((r, row_values))

    return diff_rows


# =========================
# API
# =========================
@app.post("/process")
async def process(
    base_xls: UploadFile = File(...),
    data_zip: UploadFile = File(...),
    captcha: str = Form(...)
):
    
    # 工作目錄
    work = tempfile.mkdtemp(prefix="excel_")

    # base
    base_path = os.path.join(work, base_xls.filename)

    with open(base_path, "wb") as f:
        shutil.copyfileobj(base_xls.file, f)
    # 直接複製成 result.xlsx
    result_path = os.path.join(work, "Result.xlsx")
    shutil.copy(base_path, result_path)
    
    # zip
    zip_path = os.path.join(work, data_zip.filename)

    with open(zip_path, "wb") as f:
        shutil.copyfileobj(data_zip.file, f)

    # 解壓縮
    unzip_dir = os.path.join(work, "unzipped")
    os.makedirs(unzip_dir, exist_ok=True)

    with zipfile.ZipFile(zip_path, "r") as z:
        z.extractall(unzip_dir)

    # 開啟 base
    base_wb = load_workbook(base_path)
    
    # 開啟 Result.xlsx
    result_wb = load_workbook(result_path)  
    
    # TXT 結果
    txt_path = os.path.join(work, "Result.txt")

    with open(txt_path, "w", encoding="utf-8") as txt_fp:

        # 掃描 ZIP
        for file in os.listdir(unzip_dir):

            if not file.endswith((".xlsx", ".xlsm")):
                continue

            src_path = os.path.join(unzip_dir, file)

            #print("處理:", file)

            src_wb = load_workbook(src_path, data_only=False)

            # sheet 比較
            for sheet_name in src_wb.sheetnames:

                if sheet_name not in base_wb.sheetnames:
                    continue

                base_ws = base_wb[sheet_name]
                src_ws = src_wb[sheet_name]
                result_ws = result_wb[sheet_name]	#add

                # 差異
                diff_rows = compare_sheet(base_ws, src_ws)
                
                for row_num, values in diff_rows:
                    
                    for col_offset, value in enumerate(values):

                        # 空值跳過
                        if value is None:
                            continue

                        value_str = str(value).strip()

                        # 空白跳過
                        if value_str == "":
                            continue

                        col_num = 1 + col_offset

                        # Result Cell
                        target_cell = result_ws.cell(
                            row=row_num,
                            column=col_num
                        )
                        # 只改值
                        target_cell.value = value
                           
    # 儲存
    result_wb.save(result_path)
    
    
    # 驗證
    if captcha != current_captcha:

        return Response(
            content="Captcha Error",
            status_code=400
        )
    return FileResponse(
        path=result_path,
        filename="Result.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


